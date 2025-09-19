#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M2阶段：Excel公式逆向工程和算法提取
分析 RN8211B V3校表计算.xlsx 的关键算法逻辑
"""

import openpyxl
import re
from typing import Dict, List, Tuple, Any

class ExcelFormulaExtractor:
    """Excel公式提取器"""

    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename, data_only=False)
        self.worksheet = self.workbook.active

    def extract_key_formulas(self) -> Dict[str, Any]:
        """提取关键计算公式"""
        key_cells = {
            # 根据需求文档中提到的关键单元格
            'B25': '头部字段',
            'B26': '密码域',
            'B27': '操作者码',
            'C39': 'DI原码',
            'B32': '关键计算公式',
            'D39': '数据域计算',
            'E39': '最终帧构造'
        }

        results = {}

        for cell_addr, description in key_cells.items():
            try:
                cell = self.worksheet[cell_addr]
                cell_info = {
                    'address': cell_addr,
                    'description': description,
                    'value': cell.value,
                    'formula': cell.formula if hasattr(cell, 'formula') else None,
                    'data_type': str(type(cell.value).__name__)
                }
                results[cell_addr] = cell_info
                print(f"{cell_addr} ({description}): {cell.value}")
                if hasattr(cell, 'formula') and cell.formula:
                    print(f"  公式: {cell.formula}")

            except Exception as e:
                print(f"读取 {cell_addr} 时出错: {e}")

        return results

    def scan_formulas_in_range(self, start_row: int, end_row: int,
                              start_col: int, end_col: int) -> List[Dict]:
        """扫描指定范围内的所有公式"""
        formulas = []

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if hasattr(cell, 'formula') and cell.formula:
                    formula_info = {
                        'address': f"{chr(64 + col)}{row}",
                        'formula': cell.formula,
                        'value': cell.value
                    }
                    formulas.append(formula_info)

        return formulas

    def find_frame_construction_logic(self) -> Dict[str, Any]:
        """寻找帧构造逻辑"""
        # 扫描包含十六进制数据或帧相关的单元格
        frame_patterns = [
            r'68.*68',  # DL/T645帧头模式
            r'[0-9A-F]{2}(\s+[0-9A-F]{2})+',  # 十六进制序列
            r'0x[0-9A-F]+',  # 0x格式十六进制
        ]

        frame_cells = []

        # 扫描前50行50列寻找相关数据
        for row in range(1, 51):
            for col in range(1, 51):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    for pattern in frame_patterns:
                        if re.search(pattern, str(cell.value), re.IGNORECASE):
                            frame_cells.append({
                                'address': f"{chr(64 + col)}{row}",
                                'value': cell.value,
                                'formula': getattr(cell, 'formula', None)
                            })
                            break

        return frame_cells

    def extract_calibration_parameters(self) -> Dict[str, Any]:
        """提取校准参数"""
        # 寻找可能包含校准参数的区域
        param_keywords = ['标准', '电压', '电流', '功率', '相位', '增益', 'offset']
        params = {}

        for row in range(1, 100):
            for col in range(1, 20):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    for keyword in param_keywords:
                        if keyword in str(cell.value):
                            # 检查右侧是否有数值
                            value_cell = self.worksheet.cell(row=row, column=col+1)
                            if value_cell.value is not None:
                                params[f"{chr(64 + col)}{row}"] = {
                                    'label': cell.value,
                                    'value': value_cell.value,
                                    'value_cell': f"{chr(64 + col + 1)}{row}"
                                }

        return params

def analyze_dlt645_algorithm():
    """分析DL/T645算法实现"""
    print("=== M2阶段：Excel算法逆向工程 ===\n")

    try:
        extractor = ExcelFormulaExtractor("RN8211B V3校表计算.xlsx")

        print("1. 提取关键公式:")
        print("-" * 40)
        key_formulas = extractor.extract_key_formulas()

        print("\n2. 寻找帧构造逻辑:")
        print("-" * 40)
        frame_logic = extractor.find_frame_construction_logic()
        for item in frame_logic[:10]:  # 只显示前10个
            print(f"{item['address']}: {item['value']}")

        print("\n3. 提取校准参数:")
        print("-" * 40)
        params = extractor.extract_calibration_parameters()
        for addr, info in list(params.items())[:10]:  # 只显示前10个
            print(f"{addr}: {info['label']} = {info['value']}")

        print("\n4. 扫描公式范围 (A1:J50):")
        print("-" * 40)
        formulas = extractor.scan_formulas_in_range(1, 50, 1, 10)
        for formula in formulas[:5]:  # 只显示前5个公式
            print(f"{formula['address']}: {formula['formula']} = {formula['value']}")

        # 保存分析结果供后续使用
        analysis_results = {
            'key_formulas': key_formulas,
            'frame_logic': frame_logic,
            'calibration_params': params,
            'formulas': formulas
        }

        return analysis_results

    except FileNotFoundError:
        print("错误: 找不到 RN8211B V3校表计算.xlsx 文件")
        return None
    except Exception as e:
        print(f"分析过程中出错: {e}")
        return None

if __name__ == "__main__":
    results = analyze_dlt645_algorithm()
    if results:
        print(f"\n=== 分析完成 ===")
        print(f"发现 {len(results['key_formulas'])} 个关键公式")
        print(f"发现 {len(results['frame_logic'])} 个帧相关单元格")
        print(f"发现 {len(results['calibration_params'])} 个校准参数")
        print(f"发现 {len(results['formulas'])} 个计算公式")