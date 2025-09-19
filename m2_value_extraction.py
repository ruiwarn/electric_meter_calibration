#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M2阶段：提取Excel实际计算值
获取关键公式的实际计算结果，用于算法实现
"""

import openpyxl

def extract_actual_values():
    """提取Excel中的实际计算值"""
    print("=== M2阶段：提取实际计算值 ===\n")

    try:
        # 读取计算后的值
        workbook = openpyxl.load_workbook("RN8211B V3校表计算.xlsx", data_only=True)
        ws = workbook.active

        # 关键单元格的实际值
        key_cells = [
            'B25', 'B26', 'B27', 'C34', 'C36', 'C39', 'D39', 'E39', 'B32', 'B34'
        ]

        values = {}
        print("1. 关键单元格实际值:")
        print("-" * 40)
        for cell_addr in key_cells:
            cell_value = ws[cell_addr].value
            values[cell_addr] = cell_value
            print(f"{cell_addr}: {cell_value}")

        # 寻找具体的0x33偏置计算
        print("\n2. 数据域相关值:")
        print("-" * 40)
        data_cells = ['E42', 'E43', 'E44', 'E45', 'E46', 'E47']
        for cell_addr in data_cells:
            try:
                cell_value = ws[cell_addr].value
                print(f"{cell_addr}: {cell_value}")
            except:
                print(f"{cell_addr}: (不存在)")

        return values

    except Exception as e:
        print(f"提取过程中出错: {e}")
        return None

def analyze_di_transformation():
    """分析DI变换逻辑"""
    print("\n3. DI变换逻辑分析:")
    print("-" * 40)

    # 已知的DI原码
    di_original = "00F81500"
    print(f"DI原码: {di_original}")

    # 按照D39公式进行变换: RIGHT(C39,2) MID(C39,5,2) MID(C39,3,2) LEFT(C39,2)
    # RIGHT(C39,2) = 最后2位 = "00"
    # MID(C39,5,2) = 第5位开始2位 = "15"
    # MID(C39,3,2) = 第3位开始2位 = "F8"
    # LEFT(C39,2) = 前2位 = "00"

    right_2 = di_original[-2:]  # "00"
    mid_5_2 = di_original[4:6]  # "15"
    mid_3_2 = di_original[2:4]  # "F8"
    left_2 = di_original[:2]    # "00"

    di_transformed = f"{right_2} {mid_5_2} {mid_3_2} {left_2}"
    print(f"DI变换后: {di_transformed}")

    # 这实际上是字节序的翻转: 00F81500 -> 00 15 F8 00
    bytes_original = [di_original[i:i+2] for i in range(0, len(di_original), 2)]
    bytes_reversed = bytes_original[::-1]  # 翻转字节序
    di_reversed = " ".join(bytes_reversed)
    print(f"字节序翻转: {di_reversed}")

    return {
        'original': di_original,
        'transformed': di_transformed.replace(' ', ''),
        'is_byte_reverse': di_transformed == di_reversed
    }

def analyze_offset_calculation():
    """分析0x33偏置计算"""
    print("\n4. 0x33偏置计算分析:")
    print("-" * 40)

    # 模拟数据域: 00 15 F8 00 (DI) + 4字节参数数据
    sample_data = ["00", "15", "F8", "00", "AA", "BB", "CC", "DD"]
    print(f"原始数据: {' '.join(sample_data)}")

    # 应用0x33偏置
    offset_data = []
    for hex_byte in sample_data:
        original = int(hex_byte, 16)
        offset = (original + 0x33) & 0xFF
        offset_data.append(f"{offset:02X}")

    print(f"偏置后数据: {' '.join(offset_data)}")

    return {
        'original': sample_data,
        'offset': offset_data
    }

def calculate_checksum_example():
    """计算校验和示例"""
    print("\n5. 校验和计算示例:")
    print("-" * 40)

    # 模拟完整帧数据 (从第二个68开始到数据域末尾)
    # 68 + 控制码 + 数据长度 + 数据域
    checksum_data = "68 14 08 33 48 2B 33 DD EE FF 00 AA"
    hex_bytes = checksum_data.split()

    checksum = 0
    for hex_byte in hex_bytes:
        checksum += int(hex_byte, 16)

    checksum_final = checksum & 0xFF
    print(f"校验和数据: {checksum_data}")
    print(f"校验和计算: {checksum:04X} & 0xFF = {checksum_final:02X}")

    return checksum_final

if __name__ == "__main__":
    # 提取实际值
    values = extract_actual_values()

    if values:
        # 分析各种变换
        di_analysis = analyze_di_transformation()
        offset_analysis = analyze_offset_calculation()
        checksum_example = calculate_checksum_example()

        print(f"\n=== 分析总结 ===")
        print(f"DI变换验证: {'通过' if di_analysis['is_byte_reverse'] else '需要进一步分析'}")
        print(f"完整帧格式: {values.get('B32', 'N/A')}")