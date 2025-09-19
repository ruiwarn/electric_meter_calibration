#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M2阶段：详细分析Excel关键公式
深入分析B32公式中的各个组成部分
"""

import openpyxl
import re

def analyze_frame_construction():
    """详细分析帧构造逻辑"""
    print("=== M2阶段：详细帧构造分析 ===\n")

    try:
        workbook = openpyxl.load_workbook("RN8211B V3校表计算.xlsx", data_only=False)
        ws = workbook.active

        # B32公式的组成部分
        frame_components = {
            'B25': '头部字段(起始+地址+控制)',
            'C34': '数据长度',
            'E39': '数据域(DI翻转)',
            'B26': '密码域',
            'B27': '操作者码',
            'B34': '校验和前部分',
            'C36': '校验和',
            '16': '结束符'
        }

        print("1. B32帧构造公式组成分析:")
        print("-" * 50)

        for cell_addr, description in frame_components.items():
            if cell_addr == '16':
                print(f"常量 16: {description}")
                continue

            try:
                cell = ws[cell_addr]
                print(f"{cell_addr} ({description}):")
                print(f"  值: {cell.value}")
                if hasattr(cell, 'formula') and cell.formula:
                    print(f"  公式: {cell.formula}")
                print()
            except Exception as e:
                print(f"  读取错误: {e}\n")

        # 分析DI翻转逻辑 (D39和E39)
        print("\n2. DI翻转逻辑分析:")
        print("-" * 50)
        d39_cell = ws['D39']
        e39_cell = ws['E39']
        c39_cell = ws['C39']

        print(f"C39 (DI原码): {c39_cell.value}")
        print(f"D39 (翻转逻辑): {d39_cell.formula if hasattr(d39_cell, 'formula') else 'No formula'}")
        print(f"D39 值: {d39_cell.value}")
        print(f"E39 公式类型: {type(e39_cell.value)}")

        # 分析校验和计算
        print("\n3. 校验和计算分析:")
        print("-" * 50)
        checksum_cells = ['B34', 'C36']
        for cell_addr in checksum_cells:
            cell = ws[cell_addr]
            print(f"{cell_addr}:")
            print(f"  值: {cell.value}")
            if hasattr(cell, 'formula') and cell.formula:
                print(f"  公式: {cell.formula}")
            print()

        # 分析数据长度计算
        print("\n4. 数据长度计算:")
        print("-" * 50)
        c34_cell = ws['C34']
        print(f"C34 (数据长度):")
        print(f"  值: {c34_cell.value}")
        if hasattr(c34_cell, 'formula') and c34_cell.formula:
            print(f"  公式: {c34_cell.formula}")

        # 寻找0x33偏置相关逻辑
        print("\n5. 寻找0x33偏置逻辑:")
        print("-" * 50)
        for row in range(30, 50):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if '33' in str(cell.value) or '+33' in str(cell.value):
                        print(f"{chr(64 + col)}{row}: {cell.value}")
                        if hasattr(cell, 'formula') and cell.formula:
                            print(f"  公式: {cell.formula}")

        # 分析实际数据示例
        print("\n6. 完整帧示例分析:")
        print("-" * 50)
        b32_cell = ws['B32']
        if b32_cell.value:
            frame_hex = str(b32_cell.value)
            print(f"完整帧: {frame_hex}")
            print(f"帧长度: {len(frame_hex.replace(' ', '')) // 2} 字节")

            # 尝试解析帧结构
            hex_clean = frame_hex.replace(' ', '')
            if len(hex_clean) >= 20:  # 最小DL/T645帧长度
                print("\n帧结构解析:")
                print(f"  起始符: {hex_clean[0:2]}")
                print(f"  地址域: {hex_clean[2:14]}")
                print(f"  起始符2: {hex_clean[14:16]}")
                print(f"  控制码: {hex_clean[16:18]}")
                print(f"  数据长度: {hex_clean[18:20]}")
                if len(hex_clean) > 20:
                    data_len = int(hex_clean[18:20], 16)
                    data_end = 20 + data_len * 2
                    if len(hex_clean) >= data_end + 4:
                        print(f"  数据域: {hex_clean[20:data_end]}")
                        print(f"  校验和: {hex_clean[data_end:data_end+2]}")
                        print(f"  结束符: {hex_clean[data_end+2:data_end+4]}")

    except Exception as e:
        print(f"分析过程中出错: {e}")

def extract_algorithm_patterns():
    """提取算法模式"""
    print("\n=== 算法模式提取 ===")

    patterns = {
        'DI_REVERSE': 'DI字节序翻转: D39公式',
        'DATA_OFFSET': '0x33偏置: 数据域偏移',
        'CHECKSUM': '校验和: 从第二个68到数据域末尾',
        'FRAME_CONCAT': '帧拼接: B32公式'
    }

    for pattern, desc in patterns.items():
        print(f"{pattern}: {desc}")

if __name__ == "__main__":
    analyze_frame_construction()
    extract_algorithm_patterns()